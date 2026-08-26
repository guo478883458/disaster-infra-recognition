"""
河道水位识别：水位尺检测 → 数字/刻度识别 → 水位数值输出

参考 GaoKangYu/Water-Level-Recognition-With-OCR 方案：
  - YOLO 检测水位尺上的数字区域
  - OCR 识别数字值
  - 刻度线检测辅助提高精度
  - 综合计算最终水位值

引用说明：
  - 本实现参考了 GaoKangYu 的开源方案（github.com/GaoKangYu/Water-Level-Recognition-With-OCR）
  - 采用 YOLOv8 替代原方案的 YOLOv3 进行数字检测
  - 采用 PaddleOCR 替代原方案的 LSTM-OCR 进行字符识别
  - 水位计算公式沿用原方案
"""
import os
import time
import cv2
import numpy as np
from pathlib import Path
from collections import OrderedDict

from .config import (
    MODEL_WEIGHTS_DIR, CONFIDENCE_THRESHOLD, OCR_CONFIDENCE_THRESHOLD,
    PRETRAINED_WEIGHTS, XLSX_PATH, IMAGES_DIR, get_image_path
)


class WaterLevelDetector:
    """
    河道水位识别器

    管线: 水位尺数字检测 → OCR 读数 → 刻度线检测 → 水位计算
    """

    def __init__(self, weights_path: str = None, device: str = "cpu"):
        self.device = device
        self.weights_path = weights_path or PRETRAINED_WEIGHTS
        self.digit_detector = None
        self.ocr_reader = None
        self._ground_truth = None  # 缓存真值表 {image_name: water_level_cm}
        self._load_models()

    # ==================== 模型加载 ====================

    def _load_models(self):
        """加载检测模型和 OCR 模型"""
        try:
            from ultralytics import YOLO
        except ImportError:
            raise ImportError("请安装 ultralytics: pip install ultralytics")

        if os.path.exists(self.weights_path):
            self.digit_detector = YOLO(self.weights_path)
        else:
            print(f"[WARN] 数字检测权重未找到: {self.weights_path}")
            print("       将使用 YOLOv8n 预训练模型（需微调后用于水位尺数字检测）")
            self.digit_detector = YOLO("yolov8n.pt")

        try:
            import easyocr
            self.ocr_reader = easyocr.Reader(
                ["ch_sim", "en"],
                gpu=(self.device == "cuda"),
                model_storage_directory=os.path.join(
                    os.path.expanduser("~"), ".cache", "easyocr"
                ),
                download_enabled=True,
            )
        except ImportError:
            raise ImportError(
                "请安装 easyocr: pip install easyocr"
            )

    # ==================== 真值加载 ====================

    def load_ground_truth(self) -> dict:
        """
        加载 xlsx 真值表

        xlsx 结构:
          Row 1: 标题行 (ID, Image_name, year, ...)
          Row 2: 子标题行 (col I-L: 各模拟方法名)
          Row 3+: 数据行
          Col B: Image_name, Col M: Gauge water level (cm)

        Returns:
            {image_name: water_level_cm}  # 水位值单位: cm
        """
        if self._ground_truth is not None:
            return self._ground_truth

        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        ws = wb["Total"]

        gt = {}
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            image_name = row[1]  # Col B
            gauge_level = row[12]  # Col M (0-indexed: 12)
            if image_name and gauge_level is not None:
                gt[str(image_name)] = float(gauge_level)

        wb.close()
        self._ground_truth = gt
        print(f"  [INFO] 加载真值表: {len(gt)} 条记录")
        return gt

    def get_ground_truth(self, image_name: str) -> float:
        """获取单张图像的真值水位（cm）"""
        gt = self.load_ground_truth()
        return gt.get(image_name, None)

    # ==================== 管线模块 ====================

    def detect_digit_regions(self, image: np.ndarray) -> list:
        """
        检测水位尺上的数字区域
        Returns: [{"bbox":[x1,y1,x2,y2], "confidence":float, "class_id":int}, ...]
        """
        results = self.digit_detector(image, conf=CONFIDENCE_THRESHOLD)[0]
        detections = []
        if results.boxes is not None:
            for box, conf, cls_id in zip(
                results.boxes.xyxy, results.boxes.conf, results.boxes.cls
            ):
                detections.append({
                    "bbox": box.tolist(),
                    "confidence": float(conf),
                    "class_id": int(cls_id)
                })
        detections.sort(key=lambda d: d["bbox"][1])  # 按 Y 排序
        return detections

    def read_digits_ocr(self, image: np.ndarray, bbox: list) -> tuple:
        """OCR 识别数字区域 → (text, confidence)，放大 6 倍提高精度"""
        x1, y1, x2, y2 = [int(v) for v in bbox]
        crop = image[y1:y2, x1:x2]
        if crop.size == 0:
            return "", 0.0

        # 放大 6 倍提升竖排小字识别率
        h, w = crop.shape[:2]
        crop = cv2.resize(crop, (w * 6, h * 6), interpolation=cv2.INTER_LINEAR)

        result = self.ocr_reader.readtext(crop)
        if result:
            texts = [item[1] for item in result]
            confs = [item[2] for item in result]
            return "".join(texts), float(np.mean(confs))
        return "", 0.0

    def detect_tick_marks(self, image: np.ndarray, roi: list) -> int:
        """霍夫变换检测刻度线数量"""
        x1, y1, x2, y2 = [int(v) for v in roi]
        roi_img = image[y1:y2, x1:x2]
        if roi_img.size == 0:
            return 0
        gray = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150, apertureSize=3)
        lines = cv2.HoughLinesP(edges, rho=1, theta=np.pi / 180,
                                threshold=30, minLineLength=20, maxLineGap=5)
        return len(lines) if lines is not None else 0

    def calculate_water_level(self, detections: list, ocr_results: list,
                               tick_count: int, image_height: int = None) -> dict:
        """
        计算水位值（基于最小可见数字 N）

        水位尺数字从上到下为 9,8,7,6,5...（每 0.1 米一个刻度）
        最小可见数字 N → 水位 ≈ 1.N 米档位
        可选插值：用最下方数字框底 y 与框间距 ~90px/0.1m 推算小数位
        """
        if not detections or not ocr_results:
            return {"water_level_cm": None, "confidence": 0.0,
                    "details": {"error": "未检测到数字区域"}}

        # 解析 OCR 结果，过滤 0 和两位数字（误检）
        valid_digits = []
        for det, ocr in zip(detections, ocr_results):
            try:
                text = ocr.get("text", "").strip()
                if not text:
                    continue
                # 取第一个数字字符
                digits_only = "".join(ch for ch in text if ch.isdigit())
                if not digits_only:
                    continue
                val = int(digits_only)
            except (ValueError, TypeError):
                continue
            # 过滤：只保留 1~9 的单数字
            if 1 <= val <= 9:
                valid_digits.append({
                    "bbox": det["bbox"],
                    "value": val,
                    "confidence": ocr["confidence"]
                })

        if not valid_digits:
            return {"water_level_cm": None, "confidence": 0.0,
                    "details": {"error": "无法识别有效数字"}}

        # 取最小数字 N
        min_digit = min(valid_digits, key=lambda d: d["value"])
        N = min_digit["value"]

        # 基准水位 = 1.0 + N × 0.1 米
        base_cm = 100 + N * 10

        # 可选插值：用最下方数字框底 y 推算小数位
        extra_cm = 0.0
        if image_height is not None:
            bottom_digit = max(valid_digits, key=lambda d: d["bbox"][3])  # 最下方框
            y_bottom = bottom_digit["bbox"][3]
            tick_spacing = 90.0  # px per 0.1m
            fraction = (image_height - y_bottom) / tick_spacing
            fraction = max(0.0, min(1.0, fraction))
            extra_cm = round(fraction * 10, 1)

        water_level_cm = base_cm + extra_cm

        # 置信度取平均
        confs = [d["confidence"] for d in valid_digits if d["confidence"] > 0]
        avg_conf = float(np.mean(confs)) if confs else 0.0

        return {
            "water_level_cm": round(water_level_cm, 2),
            "water_level_m": round(water_level_cm / 100, 3),
            "confidence": round(avg_conf, 4),
            "details": {
                "min_digit": N,
                "base_cm": base_cm,
                "extra_cm": extra_cm,
                "digits_detected": len(valid_digits),
                "all_digits": [d["value"] for d in valid_digits]
            }
        }

    # ==================== 推理入口 ====================

    def infer(self, image_path: str) -> dict:
        """
        对单张水位尺图像执行完整推理管线

        Returns:
            {"water_level_m": float, "water_level_cm": float,
             "confidence": float, "inference_time_ms": float,
             "details": dict, "image_name": str}
        """
        t0 = time.time()

        image = cv2.imread(image_path)
        if image is None:
            return {"water_level_m": None, "water_level_cm": None,
                    "confidence": 0.0, "inference_time_ms": 0.0,
                    "details": {"error": f"无法读取图像: {image_path}"},
                    "image_name": os.path.basename(image_path)}

        image_name = os.path.basename(image_path)

        detections = self.detect_digit_regions(image)
        ocr_results = [{"text": t, "confidence": c}
                       for t, c in (self.read_digits_ocr(image, d["bbox"])
                                    for d in detections)]

        tick_count = 0
        if len(detections) >= 2:
            roi = [detections[0]["bbox"][0], detections[0]["bbox"][3],
                   detections[-1]["bbox"][2], detections[-1]["bbox"][3] + 50]
            tick_count = self.detect_tick_marks(image, roi)

        result = self.calculate_water_level(detections, ocr_results, tick_count,
                                              image_height=image.shape[0])
        result["inference_time_ms"] = round((time.time() - t0) * 1000, 2)
        result["image_name"] = image_name

        # 附加真值（如果有）
        gt = self.get_ground_truth(image_name)
        if gt is not None:
            result["ground_truth_cm"] = gt
            if result["water_level_cm"] is not None:
                result["absolute_error_cm"] = round(
                    abs(result["water_level_cm"] - gt), 2
                )

        return result

    # ==================== 批量评估 ====================

    def evaluate(self, image_paths: list) -> dict:
        """
        批量评估，计算 MAE

        Args:
            image_paths: 图像路径列表

        Returns:
            {"mae_cm": float, "mean_time_ms": float,
             "results": [each_infer_result], "n": int}
        """
        results = []
        for p in image_paths:
            r = self.infer(p)
            results.append(r)

        errors = [r["absolute_error_cm"] for r in results
                  if r.get("absolute_error_cm") is not None]
        times = [r["inference_time_ms"] for r in results if r["inference_time_ms"] > 0]

        return {
            "mae_cm": round(float(np.mean(errors)), 2) if errors else None,
            "mean_time_ms": round(float(np.mean(times)), 2) if times else None,
            "results": results,
            "n": len(results),
            "n_with_gt": len(errors)
        }


def demo_inference():
    """快速演示：对 3 张水位尺图像进行推理"""
    detector = WaterLevelDetector()
    gt = detector.load_ground_truth()

    # 从数据目录取 3 张图像
    date_dirs = sorted(os.listdir(IMAGES_DIR))
    images = []
    for dd in date_dirs:
        img_dir = os.path.join(IMAGES_DIR, dd, "images")
        if os.path.isdir(img_dir):
            for f in sorted(os.listdir(img_dir))[:5]:
                full = os.path.join(img_dir, f)
                if any(full.lower().endswith(ext) for ext in [".jpg", ".jpeg", ".png"]):
                    images.append(full)
        if len(images) >= 3:
            break

    if not images:
        print("[ERROR] 无测试图像可用")
        return

    print(f"\n{'='*65}")
    print(f"河道水位识别推理测试（{len(images)} 张）")
    print(f"{'='*65}")

    for img_path in images:
        result = detector.infer(img_path)
        name = result["image_name"]
        print(f"  图像: {name}")
        if result["water_level_cm"] is not None:
            print(f"    水位: {result['water_level_cm']:.2f} cm ({result['water_level_m']:.3f} m)")
            print(f"    置信度: {result['confidence']:.2%}")
        else:
            print(f"    结果: {result['details'].get('error', '识别失败')}")

        if result.get("ground_truth_cm"):
            print(f"    真值: {result['ground_truth_cm']:.1f} cm")
            if result.get("absolute_error_cm"):
                print(f"    绝对误差: {result['absolute_error_cm']:.2f} cm")
        print(f"    耗时: {result['inference_time_ms']:.1f} ms\n")


if __name__ == "__main__":
    demo_inference()